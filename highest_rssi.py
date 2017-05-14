import openpyxl


def read_data():
    start_row = end_row = 4                                             # Get starting row of BSSIDs
    stop = False
    while not stop:                                                     # Get ending row of BSSIDs
        if not sheet.cell(row=end_row, column=1).value:
            stop = True
        else:
            end_row += 1
    end_row -= 1

    amount_of_bssids = end_row - start_row + 1
    print('Amount of BSSIDs: ', amount_of_bssids)
    amount_of_bssid_combs = sheet.cell(row=end_row+3, column=4).value  # Get amount of possible combinations of 2 BSSIDs
    print('Amount of combinations: ', amount_of_bssid_combs)

    begin_combs = end_row + 10
    end_combs = begin_combs + amount_of_bssid_combs - 1
    rssis = []

    for j in range(start_row, end_row+1):                               # Load RSSI values
        rssis.append(sheet.cell(row=j, column=2).value)

    read_data = [start_row, end_row, rssis, begin_combs, end_combs]
    return read_data


def write_error():
    for row in range(begin_combs, end_combs+1):
        if sheet.cell(row=row, column=2).value is bssid1 and sheet.cell(row=row, column=3).value is bssid2 \
                or sheet.cell(row=row, column=2).value is bssid2 and sheet.cell(row=row, column=3).value is bssid1:
            error = sheet.cell(row=row, column=7).value
            if error is None:
                error = "No match"
            sheet_write.cell(row=location+1, column=3).value = error


book_read = openpyxl.load_workbook(filename='data\\data_locationAPI_without_RSSI.xlsx')    # Load Excel sheets
book_write = openpyxl.load_workbook(filename='data\\data_comparison.xlsx')
sheet_write = book_write.get_sheet_by_name('Highest RSSI')

for location in range(1, 37):                                            # Load a template sheet for all 36 locations
    sheet = book_read.get_sheet_by_name('BAP' + str(location))

    [start_row, end_row, rssis, begin_combs, end_combs] = read_data()  # Read combination errors

    index_max1 = rssis.index(max(rssis))                                # Find rows of data with 2 highest RSSI values
    rssis.remove(max(rssis))
    index_max2 = rssis.index(max(rssis))
    if index_max2 > index_max1:
        index_max2 += 1
    index_max1 += 4
    index_max2 += 4

    bssid1 = sheet.cell(row=index_max1, column=1).value                 # Find 2 BSSIDs with highest RSSI values
    bssid2 = sheet.cell(row=index_max2, column=1).value

    write_error()                                    # Write error of combination with the 2 BSSIDs to comparison table

    book_write.save('data\\data_comparison.xlsx')                       # Save output file
    print('BAP' + str(location) + ' saved.\n')



