from itertools import combinations
import openpyxl
import unwiredlabs
from gmplot import gmplot
from haversine import haversine


def get_data():
    start_row = end_row = 4                                     # Get starting and ending row
    stop = False
    while not stop:
        if not sheet.cell(row=end_row, column=1).value:
            stop = True
        else:
            end_row += 1
    end_row -= 1

    amount_of_bssids = end_row - start_row + 1                  # Count amount of BSSIDs
    print('Amount of BSSIDs: ' + str(amount_of_bssids))

    bssids = []                                                 # Get BSSIDs and RSSIs
    rssis = []
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=1):
        for cell in row:
            bssids.append(cell.value)
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=2, max_col=2):
        for cell in row:
            percentage = cell.value
            if percentage <= 0:                                 # Convert signal strength percentage to dBm
                dbm = -100
            elif percentage >= 100:
                dbm = -50
            else:
                dbm = (percentage / 2) - 100
            rssis.append(dbm)

    print(bssids)
    print(rssis)
    data = [bssids, rssis, start_row, end_row, amount_of_bssids]
    return data


def write_combinations_to_file():
    sheet.cell(row=end_row+2, column=4).value = amount_of_bssids            # Write amount of BSSIDS and combinations
    sheet.cell(row=end_row+3, column=4).value = len(bssid_combs)            # under the first table

    row_index = end_row + 9

    for i in range(0, len(bssid_combs)):
        sheet.cell(row=row_index + i, column=1).value = i + 1               # Write combination number in column A
        sheet.cell(row=row_index + i, column=2).value = bssid_combs[i][0]   # Write BSSID 1 in column B
        sheet.cell(row=row_index + i, column=3).value = rssi_combs[i][0]    # Write RSSI 1 in column C
        sheet.cell(row=row_index + i, column=4).value = bssid_combs[i][1]   # Write BSSID 2 in column D
        sheet.cell(row=row_index + i, column=5).value = rssi_combs[i][1]    # Write RSSI 2 in column E


def perform_request_of_combinations():
    row_index = end_row + 9
    for j in range(0, len(bssid_combs)):
        request = unwiredlabs.UnwiredRequest()
        request.addAccessPoint(bssid_combs[j][0], rssi_combs[j][0])         # Add first AP
        request.addAccessPoint(bssid_combs[j][1], rssi_combs[j][1])         # Add second AP
        connection = unwiredlabs.UnwiredConnection(key='99c1d8b93a7f37')    # API key of account 'Thomas Janssen'
        response = connection.performRequest(request)                       # Perform request

        if response.status != 'Ok':
            print(j+1, 'Error:', response.status)
            mean_latitudes.append(None)  # no matches found
            mean_longitudes.append(None)
        else:
            print(j+1, 'Response: ', response.data)
            sheet.cell(row=row_index+j, column=6).value = response.lat      # Write mean latitude of 2 BSSIDs in col F
            sheet.cell(row=row_index+j, column=7).value = response.lon      # Write mean longitude of 2 BSSIDs in col G
            sheet.cell(row=row_index+j, column=8).value = response.data['accuracy']  # Write accuracy (in m) in col H
            mean_latitudes.append(response.lat)
            mean_longitudes.append(response.lon)
            accuracies.append(response.data['accuracy'])
        book.save(file)  # save file after each line! (in case of socket error -> data saved!)


def calc_error():
    gps_latitude = sheet['H2'].value                        # Get GPS coordinate
    gps_longitude = sheet['I2'].value
    gps_coordinate = (gps_latitude, gps_longitude)

    # Write error in column I, using the haversine function
    row_index = end_row + 9
    for m in range(0, len(bssid_combs)):
        if mean_latitudes[m] and mean_longitudes[m]:  # if there is a mean, calculate the error
            mean_coordinate = (mean_latitudes[m], mean_longitudes[m])
            distance = haversine(gps_coordinate, mean_coordinate)
            sheet.cell(row=row_index+m, column=9).value = distance

    # Write mean error of all pairs of BSSIDs above the table in column D
    sum = 0
    amount = 0
    for m in range(0, len(bssid_combs)):
        if sheet.cell(row=row_index+m, column=9).value:
            sum += sheet.cell(row=row_index+m, column=9).value
            amount += 1
    mean_error = float(sum / amount)
    sheet.cell(row=end_row+5, column=4).value = mean_error

    # Calculate the mean accuracy of all combinations of 2 BSSIDs and write it to Excel
    sum = 0
    amount = 0
    for m in range(0, len(bssid_combs)):
        if sheet.cell(row=row_index+m, column=8).value:
            sum += sheet.cell(row=row_index+m, column=8).value
            amount += 1
    mean_accuracy = float(sum/amount)
    sheet.cell(row=end_row+4, column=4).value = mean_accuracy


def calc_chance():
    both_not_found = 0
    for i in range(0, len(bssid_combs)):
        if not mean_latitudes[i] and not mean_longitudes[i]:
            both_not_found += 1
    chance_both_not_found = float(both_not_found / len(bssid_combs) * 100)  # percentage both BSSIDs in a pair not found
    print('The chance of having of pair of BSSIDs both not found in the database is %.2f %%' % chance_both_not_found)
    sheet.cell(row=end_row+6, column=4).value = chance_both_not_found       # write chance of no match to Excel


def show_map():
    gmap = gmplot.GoogleMapPlotter(51.212480, 4.414351, 12)     # Create Google Maps map plotter

    map_latitudes = []                                          # Plot heatmap of mean coordinates of a combination
    map_longitudes = []
    for lat in mean_latitudes:
        if lat:
            map_latitudes.append(lat)
    for long in mean_longitudes:
        if long:
            map_longitudes.append(long)
    gmap.heatmap(map_latitudes, map_longitudes)

    gps_lat = sheet['H2'].value                                 # Plot the GPS coordinate of the location (black spot)
    gps_long = sheet['I2'].value
    gmap.circle(gps_lat, gps_long, 1, "k", ew=2)

    gmap.draw("maps\\maps_locationAPI_with_RSSI\\BAP" + str(location) + ".html")            # Save the HTML file


#  Main program
#  --------------------------------------------------------------------------------------------------------------------
file = 'data\\data_locationAPI_with_RSSI.xlsx'                                  # Load Excel sheet of a location (e.g. BAP1)
book = openpyxl.load_workbook(filename=file)

for location in range(1, 37):                                   # Load a template sheet for all 36 locations
    sheet = book.get_sheet_by_name('BAP' + str(location))

    [bssids, rssis, start_row, end_row, amount_of_bssids] = get_data()  # retrieve collected data

    bssid_combs = list(combinations(bssids, 2))                 # Get all possible combinations of 2 BSSIDs
    rssi_combs = list(combinations(rssis, 2))                   # Get all the combinations of the 2 RSSIs (in dBm)

    write_combinations_to_file()                                # Write all possible combinations to Excel file

    mean_latitudes = []
    mean_longitudes = []
    accuracies = []

    perform_request_of_combinations()                           # Request the mean coordinate of each pair of BSSIDs

    calc_error()                                                # Calculate distance between GPS and mean coordinate
    calc_chance()                                               # Calculate the chance of no match in a combination
    show_map()                                                  # Save heatmap of requested coordinates

    book.save(file)                                             # Save the data
    print('Sheet and map saved: BAP' + str(location) + '\n')

# ---------------------------------------------------------------------------------------------------------------------

