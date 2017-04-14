from itertools import combinations
import openpyxl
import unwiredlabs


def get_data():
    start_row = end_row = 4                                     # Get starting and ending row
    stop = False
    while not stop:
        if not sheet.cell(row=end_row, column=1).value:
            stop = True
        else:
            end_row += 1
    end_row -= 1

    amount_of_bssids = end_row - start_row + 1                  # Count amount of BSSIDs
    print('Amount of BSSIDs: ' + str(amount_of_bssids))

    bssids = []                                                 # Get BSSIDs and RSSIs
    rssis = []
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=1):
        for cell in row:
            bssids.append(cell.value)
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=2, max_col=2):
        for cell in row:
            percentage = cell.value
            if percentage <= 0:                                 # Convert signal strength percentage to dBm
                dbm = -100
            elif percentage >= 100:
                dbm = -50
            else:
                dbm = (percentage / 2) - 100
            rssis.append(dbm)

    print(bssids)
    print(rssis)
    data = [bssids, rssis, start_row, end_row, amount_of_bssids]
    return data


def write_combinations_to_file():
    sheet.cell(row=end_row+2, column=4).value = amount_of_bssids            # Write amount of BSSIDS and combinations
    sheet.cell(row=end_row+3, column=4).value = len(bssid_combs)            # under the first table

    row_index = end_row + 8

    for i in range(0, len(bssid_combs)):                                    # Write combination number in column A
        sheet.cell(row=row_index + i, column=1).value = i + 1

    for i in range(0, len(bssid_combs)):                                    # Write BSSID 1 in column B
        sheet.cell(row=row_index + i, column=2).value = bssid_combs[i][0]

    for i in range(0, len(bssid_combs)):                                    # Write RSSI 1 in column C
        sheet.cell(row=row_index + i, column=3).value = rssi_combs[i][0]

    for i in range(0, len(bssid_combs)):                                    # Write BSSID 2 in column D
        sheet.cell(row=row_index + i, column=4).value = bssid_combs[i][1]

    for i in range(0, len(bssid_combs)):                                    # Write RSSI 2 in column E
        sheet.cell(row=row_index + i, column=5).value = rssi_combs[i][1]


def perform_request_of_combinations():
    row_index = end_row + 8
    for j in range(0, len(bssid_combs)):
        request = unwiredlabs.UnwiredRequest()
        request.addAccessPoint(bssid_combs[j][0], rssi_combs[j][0])         # Add AP1
        request.addAccessPoint(bssid_combs[j][1], rssi_combs[j][1])         # Add AP2
        connection = unwiredlabs.UnwiredConnection(key='99c1d8b93a7f37')    # API key of account 'Thomas Janssen'
        response = connection.performRequest(request)                       # Perform request

        if response.status != 'Ok':
            print('Error:', response.status)
        else:
            print('Response: ', response.data)
            sheet.cell(row=row_index+j, column=6).value = response.lat      # Write mean latitude of 2 BSSIDs in col F
            sheet.cell(row=row_index+j, column=7).value = response.lon      # Wirte mean longitude of 2 BSSIDs in col G
            sheet.cell(row=row_index+j, column=8).value = response.data['accuracy']  # Write accuracy (in m) in col H



#  Main program
#  --------------------------------------------------------------------------------------------------------------------
file = 'data_locationAPI.xlsx'                                  # Load Excel sheet of a location (e.g. BAP1)
book = openpyxl.load_workbook(filename=file)

for location in range(1, 2):                                   # Load a template sheet for all 36 locations
    sheet = book.get_sheet_by_name('BAP' + str(location))

    [bssids, rssis, start_row, end_row, amount_of_bssids] = get_data()  # retrieve collected data

    bssid_combs = list(combinations(bssids, 2))                 # Get all possible combinations of 2 BSSIDs
    rssi_combs = list(combinations(rssis, 2))                   # Get all the combinations of the 2 RSSIs (in dBm)

    write_combinations_to_file()                                # Write all possible combinations to Excel file

    mean_latitudes = []
    mean_longitudes = []
    accuracies = []

    # perform_request_of_combinations()

    # print(mean_latitudes)
    # print(mean_longitudes)
    # calc_mean()                                                 # Calculate the mean coordinate of each pair of BSSIDs
    # calc_error()                                                # Calculate distance between GPS and mean coordinate
    # calc_chances()                                              # Calculate the chance a BSSID is not found in database
    # show_map()                                                  # Save heatmap of requested coordinates

    book.save(file)                                             # Save the data
    print('Sheet and map saved: BAP' + str(location) + '\n')

# ---------------------------------------------------------------------------------------------------------------------

