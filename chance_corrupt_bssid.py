import openpyxl


def get_data():
    start_row = end_row = 4                                 # Get starting row of BSSIDs
    stop = False
    while not stop:                                         # Get ending row of BSSIDs
        if not sheet.cell(row=end_row, column=1).value:
            stop = True
        else:
            end_row += 1
    end_row -= 1

    amount_of_bssids = end_row - start_row + 1
    print('Amount of BSSIDs: ', amount_of_bssids)

    amount_of_bssid_combs = sheet.cell(row=end_row+3, column=4).value  # Get amount of possible combinations of 2 BSSIDs
    print('Amount of combinations: ', amount_of_bssid_combs)  # @TODO change for LocationAPI

    begin_combs = end_row + 10  # @TODO change column for LocationAPI
    end_combs = begin_combs + amount_of_bssid_combs - 1
    errors = []

    for i in range(begin_combs, end_combs+1):
        if sheet.cell(row=i, column=7).value:  # @TODO change column for LocationAPI
            errors.append(sheet.cell(row=i, column=7).value)       # Load distance errors (without None)# @TODO change column for LocationAPI
    read_data = [begin_combs, end_combs, errors]
    return read_data


book_read = openpyxl.load_workbook(filename='data\\data_locationAPI_without_RSSI.xlsx')        # Load Excel sheet with all location errors # @TODO change for LocationAPI
book_write = openpyxl.load_workbook(filename='data\\data_comparison.xlsx')
sheet_write = book_write.get_sheet_by_name('Chance of corrupt BSSID')

for location in range(1, 37):                                           # Load a template sheet for all 36 locations
    sheet = book_read.get_sheet_by_name('BAP' + str(location))

    [begin_combs, end_combs, errors] = get_data()                       # read combination errors
    amount__of_corrupt_errors = 0
    for error in errors:
        if error > 0.5:                                                 # threshold = 500 meters
            amount__of_corrupt_errors += 1
    chance = amount__of_corrupt_errors / len(errors) * 100              # Calculate the chance of error above threshold
    print('The chance of a corrupt bssid is %.2f' % chance, '%')
    sheet_write.cell(row=location+1, column=3).value = chance               # @TODO change column for LocationAPI

    book_write.save('data\\data_comparison.xlsx')
    print('BAP' + str(location) + ' saved.\n')



