"""
Plot a map with Antwerp in the center. The radius of each circle represents the mean/median error per location.
"""

import openpyxl
import gmplot

errors_wigle = []
errors_locationAPI = []
gps_latitudes = []
gps_longitudes = []

book1 = openpyxl.load_workbook(filename='data_comparison.xlsx')     # Load Excel sheet with all location errors
sheet1 = book1.get_sheet_by_name('Mean + median errors')
for i in range(2, 38):
    errors_wigle.append(sheet1.cell(row=i, column=4).value)         # Load WiGLE and locationAPI errors
    errors_locationAPI.append(sheet1.cell(row=i, column=5).value)

book2 = openpyxl.load_workbook(filename='data_wigle.xlsx')          # Load GPS coordinates of each location
for i in range(1, 37):
    sheet2 = book2.get_sheet_by_name('BAP'+str(i))
    gps_latitudes.append(sheet2['J2'].value)
    gps_longitudes.append(sheet2['K2'].value)

gmap = gmplot.GoogleMapPlotter(51.212480, 4.414351, 12)             # Create Google Maps map plotter

for i in range(0, len(errors_wigle)):                               # Plot WiGLE circles in blue
    gmap.circle(gps_latitudes[i], gps_longitudes[i], errors_wigle[i]*1000, "blue", ew=2)
for i in range(0, len(errors_locationAPI)):                         # Plot LocationAPI circles in orange
    gmap.circle(gps_latitudes[i], gps_longitudes[i], errors_locationAPI[i]*1000, "orange", ew=2)

gmap.draw("BAP_median_errors_comparison.html")                      # Save HTML file

