import openpyxl
import gmplot
from haversine import haversine
from itertools import combinations


def get_data():
    start_row = end_row = 4                                 # Get starting row

    stop = False
    while not stop:                                         # Get ending row
        if not sheet.cell(row=end_row, column=1).value:
            stop = True
        else:
            end_row += 1
    end_row -= 1

    amount_of_bssids = end_row - start_row + 1
    print('Amount of BSSIDs: ' + str(amount_of_bssids))

    latitudes = []
    longitudes = []
    bssids = []

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=1):   # put BSSIDs in list
        for cell in row:
            bssids.append(cell.value)
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=3, max_col=3):   # put latitudes in list
        for cell in row:
            latitudes.append(cell.value)
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=4, max_col=4):   # put longitudes in list
        for cell in row:
            longitudes.append(cell.value)

    read_data = [start_row, end_row, amount_of_bssids, bssids, latitudes, longitudes]
    return read_data


def write_combinations_to_file():

    # Write amount of BSSIDS and combinations under the first table
    sheet.cell(row=end_row + 2, column=3).value = amount_of_bssids
    sheet.cell(row=end_row + 3, column=3).value = len(bssid_combs)

    row_index = end_row + 6

    for i in range(0, len(bssid_combs)):
        sheet.cell(row=row_index + i, column=1).value = i + 1               # Write combination number in column A
        sheet.cell(row=row_index + i, column=2).value = bssid_combs[i][0]   # Write BSSID 1 in column B
        sheet.cell(row=row_index + i, column=3).value = bssid_combs[i][1]   # Write BSSID 2 in column C
        sheet.cell(row=row_index + i, column=4).value = lat_combs[i][0]     # Write latitude of BSSID 1 in column D
        sheet.cell(row=row_index + i, column=5).value = long_combs[i][0]    # Write longitude of BSSID 1 in column E
        sheet.cell(row=row_index + i, column=6).value = lat_combs[i][1]     # Write latitude of BSSID 2 in column F
        sheet.cell(row=row_index + i, column=7).value = long_combs[i][1]    # Write longitude of BSSID 2 in column G


def calc_mean():
    # Write mean latitude of the 2 BSSIDs in column H
    row_index = end_row + 6
    for i in range(0, len(bssid_combs)):
        if not lat_combs[i][0] and not lat_combs[i][1]:  # both BSSIDs were not found in database
            mean_latitude = None
        elif not lat_combs[i][0]:  # one of two BSSIDs was not found in database
            mean_latitude = lat_combs[i][1]
        elif not lat_combs[i][1]:
            mean_latitude = lat_combs[i][0]
        else:
            mean_latitude = (lat_combs[i][0] + lat_combs[i][1]) / 2

        sheet.cell(row=row_index+i, column=8).value = mean_latitude
        mean_latitudes.append(mean_latitude)

    # Write mean longitude of the 2 BSSIDs in column I
    row_index = end_row + 6
    for i in range(0, len(bssid_combs)):
        if not long_combs[i][0] and not long_combs[i][1]:   # both BSSIDs were not found in database
            mean_longitude = None
        elif not long_combs[i][0]:                          # one of two BSSIDs was not found in database
            mean_longitude = long_combs[i][1]
        elif not long_combs[i][1]:
            mean_longitude = long_combs[i][0]
        else:
            mean_longitude = (long_combs[i][0] + long_combs[i][1]) / 2

        sheet.cell(row=row_index + i, column=9).value = mean_longitude
        mean_longitudes.append(mean_longitude)


def calc_error():
    gps_latitude = sheet['J2'].value                                    # Get GPS coordinate
    gps_longitude = sheet['K2'].value
    gps_coordinate = (gps_latitude, gps_longitude)

    # Write error in column J, using the haversine function
    row_index = end_row + 6
    for i in range(0, len(bssid_combs)):
        if mean_latitudes[i] and mean_longitudes[i]:                    # if there is a mean, calculate the error
            mean_coordinate = (mean_latitudes[i], mean_longitudes[i])
            distance = haversine(gps_coordinate, mean_coordinate)
            sheet.cell(row=row_index+i, column=10).value = distance

    # Write mean error of all pairs of BSSIDs above the table in column J
    sum = 0
    amount = 0
    for i in range(0, len(bssid_combs)):
        if sheet.cell(row=row_index+i, column=10).value:
            sum += sheet.cell(row=row_index+i, column=10).value
            amount += 1
    mean_error = float(sum / amount)
    sheet.cell(row=end_row+3, column=10).value = mean_error


def calc_chances():
    bssids_not_found = 0
    for i in range(0, len(bssids)):
        if not latitudes[i] and not longitudes[i]:                          # if no match in database
            bssids_not_found += 1
    chance_not_found = float(bssids_not_found / amount_of_bssids * 100)     # percentage of total BSSIDs not found
    print('%.2f %% of BSSIDs was not found in the Wigle database.' % chance_not_found)

    both_not_found = 0
    for i in range(0, len(bssid_combs)):
        if not lat_combs[i][0] and not lat_combs[i][1]:
            both_not_found += 1
    chance_both_not_found = float(both_not_found / len(bssid_combs) * 100)  # percentage both BSSIDs in a pair not found
    print('The chance of having of pair of BSSIDs both not found in the database is %.2f %%' % chance_both_not_found)

    sheet.cell(row=end_row+2, column=8).value = chance_not_found            # write chances to Excel
    sheet.cell(row=end_row+3, column=8).value = chance_both_not_found


def show_map():
    gmap = gmplot.GoogleMapPlotter(51.212480, 4.414351, 12)     # Create Google Maps map plotter

    map_latitudes = []                                          # Plot heatmap of the found request coordinates
    map_longitudes = []
    for lat in latitudes:
        if lat:
            map_latitudes.append(lat)
    for long in longitudes:
        if long:
            map_longitudes.append(long)
    gmap.heatmap(map_latitudes, map_longitudes)

    gps_lat = sheet['J2'].value                                 # Plot the GPS coordinate of the location (black spot)
    gps_long = sheet['K2'].value
    gmap.circle(gps_lat, gps_long, 1, "k", ew=2)

    gmap.draw("maps_wigle\\BAP" + str(location) + ".html")      # Save the HTML file


#  Main program
#  --------------------------------------------------------------------------------------------------------------------
file = 'data_wigle.xlsx'                                        # Load Excel sheet of a location (e.g. BAP1)
book = openpyxl.load_workbook(filename=file)

for location in range(1, 37):                                   # Load a template sheet for all 36 locations
    sheet = book.get_sheet_by_name('BAP' + str(location))

    [start_row, end_row, amount_of_bssids, bssids, latitudes, longitudes] = get_data()  # retrieve requested data

    bssid_combs = list(combinations(bssids, 2))                 # Get all possible combinations of 2 BSSIDs
    lat_combs = list(combinations(latitudes, 2))                # Get all the combinations of the measured latitudes
    long_combs = list(combinations(longitudes, 2))              # Get all the combinations of the measured longitudes

    write_combinations_to_file()                                # Write all possible combinations to Excel file

    mean_latitudes = []
    mean_longitudes = []

    calc_mean()                                                 # Calculate the mean coordinate of each pair of BSSIDs
    calc_error()                                                # Calculate distance between GPS and mean coordinate
    calc_chances()                                              # Calculate the chance a BSSID is not found in database
    show_map()                                                  # Save heatmap of requested coordinates

    book.save(file)                                             # Save the data
    print('Sheet and map saved: BAP' + str(location) + '\n')
# ---------------------------------------------------------------------------------------------------------------------








